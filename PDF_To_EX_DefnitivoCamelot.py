# Converter PDF para excel usando Python. Deixei de usar o tabulas trocando por Camelot (que usa também o ghostScript)
# Coloquei tudo de Log em 03/06/25
import pandas as pd
import camelot  # Substituindo tabula por camelot
import tkinter as tk 
from tkinter import filedialog, messagebox, ttk
import logging

# Para fazer o logg funcionar temos que por essa estrutura no início do código
logging.basicConfig(level=logging.INFO)
logg = logging.getLogger(__name__)

################ Barra de Progresso Funções ################

def start_progress():
    logg.info("Iniciando barra de progresso...")
    # Exibe a barra de progresso e inicia animação
    progressbar.pack(pady=10)
    progressbar.start(10)  # Inicia animação da barra indeterminada (velocidade 10ms)
    # Desabilita o botão para evitar múltiplos cliques durante o processamento
    botao.config(state='disabled')
    # Atualiza a interface para mostrar mudanças imediatamente
    janela.update_idletasks()

def stop_progress():
    logg.info("Parando barra de progresso...")
    # Para a animação da barra de progresso e a oculta
    progressbar.stop()
    progressbar.pack_forget()
    # Reabilita o botão após o processamento
    botao.config(state='normal')
    # Atualiza a interface para refletir as mudanças
    janela.update_idletasks()

################ Pegando Documento da Interface + LENDO O PDF ################

def processamento():

    start_progress()
    logg.info("Processamento iniciado.")

    ################ Pegando Documento da Interface  ################

    # Abre janela para o user pegar arquivo em PDF
    caminho_arquivo = filedialog.askopenfilename(  # filedialog: Permite abrir janelas para o usuário selecionar arquivos ou escolher onde salvar.
        title="Selecione arquivo em PDF: ",
        filetypes=[("Arquivo PDF", "*.pdf")]
    )

    if caminho_arquivo:
        logg.info(f"Arquivo PDF selecionado: {caminho_arquivo}")
    # Se o usuário não escolher nada exibe esse aviso e encerra o processo 
    if not caminho_arquivo:
         logg.warning("Nenhum arquivo PDF foi selecionado.")
         messagebox.showwarning("Aviso", "Nenhum arquivo foi selecionado.")  # messagebox: Permite exibir caixas de mensagem (avisos, erros, informações)
         stop_progress()
         return

    ################ LENDO O PDF  ################
    try:
        logg.info("Lendo o PDF...")

        # Tenta extrair tabelas do PDF usando o método lattice (funciona melhor para PDFs que possuem bordas visíveis nas tabelas)
        tables_lattice = camelot.read_pdf(
            caminho_arquivo,
            pages='all',      # Analisa todas as páginas do PDF
            flavor='lattice'  # Método de extração baseado em linhas e bordas
        )
        logg.info(f"Total de tabelas encontradas com 'lattice': {tables_lattice.n}")

        # Tenta extrair tabelas do PDF usando o método stream (funciona melhor para PDFs com tabelas sem bordas visíveis, baseando-se em espaços)
        tables_stream = camelot.read_pdf(
            caminho_arquivo,
            pages='all',      # Analisa todas as páginas do PDF
            flavor='stream'   # Método de extração baseado em posicionamento do texto
        )
        logg.info(f"Total de tabelas encontradas com 'stream': {tables_stream.n}")

        # Compara o número de tabelas encontradas por cada método para escolher o melhor
        if tables_lattice.n >= tables_stream.n and tables_lattice.n != 0:
            tables = tables_lattice
            metodo = 'lattice'
        elif tables_stream.n != 0:
            tables = tables_stream
            metodo = 'stream'
        else:
            logg.warning("Nenhuma tabela encontrada no PDF.")
            messagebox.showinfo("Informação", "Nenhuma tabela foi encontrada no PDF.")
            stop_progress()
            return

        logg.info(f"Modo usado para leitura: {metodo}")

        # Combina todas as tabelas em um único DataFrame usando o pd.concat - é uma função do Pandas para unir DataFrames.
        dataframe_complete = pd.concat([t.df for t in tables])   
        logg.info("Tabelas combinadas em um único DataFrame.")

        # Abre a janela para salvar o arquivo, forçando extensão .xlsx
        caminho_saida = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx")],
            title="Salvar arquivo como:"
        )

        if caminho_saida:
            logg.info(f"Caminho de saída selecionado: {caminho_saida}")
        else:
            logg.warning("Usuário cancelou a escolha do caminho de saída.")

        ###################### Tratamento de Linhas e colunas vazias ######################

        logg.info("Iniciando tratamento do dataframe...")

        # Remove espaços extras em cada célula do DataFrame, apenas nas strings
        dataframe_complete = dataframe_complete.applymap(lambda x: x.strip() if isinstance(x, str) else x)  
        # Remove colunas que estão totalmente vazias (NaN)
        dataframe_complete = dataframe_complete.dropna(how='all', axis=1)  
        # Remove linhas que estão totalmente vazias (NaN)
        dataframe_complete = dataframe_complete.dropna(how='all', axis=0)  

        logg.info("Tratamento do dataframe concluído.")

        ###################### Exportando com openpyxl formatado ######################
        if caminho_saida:
            logg.info("Iniciando exportação para Excel...")

            from openpyxl import Workbook
            from openpyxl.styles import Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.utils.dataframe import dataframe_to_rows

            wb = Workbook()
            ws = wb.active
            ws.title = "PDF_Convertido"

            for r in dataframe_to_rows(dataframe_complete, index=False, header=True):
                ws.append(r)

            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    if cell.value is not None:
                        cell.border = border

            for col in range(ws.max_column + 1, 50):
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].hidden = True

            for row in range(ws.max_row + 1, 200):
                ws.row_dimensions[row].hidden = True

            wb.save(caminho_saida)
            logg.info(f"Arquivo Excel salvo com sucesso em: {caminho_saida}")
            messagebox.showinfo("Sucesso", f"Arquivo salvo com sucesso em:\n{caminho_saida}")
        
        else:
            logg.warning("Operação cancelada pelo usuário no momento de salvar.")
            messagebox.showwarning("Cancelado", "Operação cancelada pelo usuário.")

    except PermissionError:
        logg.error("Permissão negada: O arquivo Excel de destino está em aberto.")
        messagebox.showerror("Permissão negada", "Você deve estar mantendo o arquivo Excel de destino em aberto.")    
    
    except Exception as e:
        logg.exception("Erro inesperado durante o processamento.")
        messagebox.showerror("Erro", f"Erro ao processar o arquivo:\n{e}")

    stop_progress()
    logg.info("Processamento finalizado.")

################ Tkinter Básico ################

# Cria janela principal da aplicação
janela = tk.Tk()
janela.geometry("350x175")  # Define tamanho da janela
janela.title("Conversor Pdf powered by Mobi")  # Título da janela

# Título maior e em negrito
tk.Label(janela, text="Conversor de PDF para Excel", font=("Helvetica", 13, "bold")).pack(pady=(10, 15))

# Botão para iniciar a conversão chamando a função processamento
botao = tk.Button(janela, 
                  text="Converter PDF", 
                  command=processamento, 
                  font=("Helvetica", 13, "bold"),
                  height=1, 
                  width=15)
botao.pack(pady=1)

# Barra de progresso indeterminada (não mostra porcentagem)
progressbar = ttk.Progressbar(janela, orient='horizontal', length=300, mode='indeterminate')
progressbar.pack_forget()  # Inicialmente oculta a barra

# Executa a interface gráfica
janela.mainloop()