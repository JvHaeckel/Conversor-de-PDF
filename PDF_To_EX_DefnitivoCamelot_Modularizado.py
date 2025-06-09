# Converter PDF para Excel usando Python. Deixei de usar o tabulas trocando por Camelot (que usa também o ghostScript)
# Mesmo código do PDF_To_EX_DefnitivoCamelot.py sendo que modularizado em 03/06/25

import pandas as pd
import camelot  # Substituindo tabula por camelot
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import logging

# Para fazer o logg funcionar temos que por essa estrutura no início do código
logging.basicConfig(level=logging.INFO)
logg = logging.getLogger(__name__)

################ Barra de Progresso Funções ################

def start_progress():
    progressbar.pack(pady=10)
    progressbar.start(10)  # Inicia animação da barra indeterminada (velocidade 10ms)
    botao.config(state='disabled')
    janela.update_idletasks()

def stop_progress():
    progressbar.stop()
    progressbar.pack_forget()
    botao.config(state='normal')
    janela.update_idletasks()

################## Funções Separadas #######################

def selecionar_arquivo_pdf():
    # Abre janela para o user pegar arquivo em PDF
    caminho = filedialog.askopenfilename(
        title="Selecione arquivo em PDF:",
        filetypes=[("Arquivo PDF", "*.pdf")]
    )
    return caminho


def selecionar_caminho_saida():
    # Abre janela para salvar o arquivo Excel
    caminho = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Arquivos Excel", "*.xlsx")],
        title="Salvar arquivo como:"
    )
    return caminho


def extrair_tabelas_pdf(caminho_arquivo):
    ################ LENDO O PDF ################

    # Extrai tabelas usando o método lattice (bordas visíveis)
    tables_lattice = camelot.read_pdf(
        caminho_arquivo,
        pages='all',
        flavor='lattice'
    )

    # Extrai tabelas usando o método stream (sem bordas, baseado em espaços)
    tables_stream = camelot.read_pdf(
        caminho_arquivo,
        pages='all',
        flavor='stream'
    )

    # Escolhe o melhor método
    if tables_lattice.n >= tables_stream.n and tables_lattice.n != 0:
        tables = tables_lattice
        metodo = 'lattice'
    elif tables_stream.n != 0:
        tables = tables_stream
        metodo = 'stream'
    else:
        return None, None  # Nenhuma tabela encontrada

    logg.info(f"Modo usado para leitura: {metodo}")

    # Combina todas as tabelas em um único DataFrame
    dataframe_complete = pd.concat([t.df for t in tables])

    return dataframe_complete, metodo


def tratar_dataframe(df):
    ################ Tratamento de Linhas e Colunas Vazias ################

    # Remove espaços extras em cada célula
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    # Remove colunas totalmente vazias
    df = df.dropna(how='all', axis=1)
    # Remove linhas totalmente vazias
    df = df.dropna(how='all', axis=0)

    return df


def exportar_para_excel(df, caminho_saida):
    ################ Exportando com openpyxl formatado ################

    from openpyxl import Workbook
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    ws = wb.active
    ws.title = "PDF_Convertido"

    # Adiciona os dados do DataFrame na planilha
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Define borda fina
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Aplica borda nas células preenchidas
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                cell.border = border

    # Oculta colunas extras após os dados
    for col in range(ws.max_column + 1, 50):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].hidden = True

    # Oculta linhas extras após os dados
    for row in range(ws.max_row + 1, 200):
        ws.row_dimensions[row].hidden = True

    # Salva o arquivo Excel no local escolhido pelo usuário
    wb.save(caminho_saida)


################## Função Principal de Processamento #######################

def processamento():
    start_progress()

    try:
        ################ Pegando Documento da Interface ################
        caminho_arquivo = selecionar_arquivo_pdf()

        if not caminho_arquivo:
            messagebox.showwarning("Aviso", "Nenhum arquivo foi selecionado.")
            stop_progress()
            return

        ################ Extração das Tabelas ################
        dataframe, metodo = extrair_tabelas_pdf(caminho_arquivo)

        if dataframe is None:
            messagebox.showinfo("Informação", "Nenhuma tabela foi encontrada no PDF.")
            stop_progress()
            return

        ################ Tratamento ################
        dataframe = tratar_dataframe(dataframe)

        ################ Salvamento ################
        caminho_saida = selecionar_caminho_saida()

        if caminho_saida:
            exportar_para_excel(dataframe, caminho_saida)
            messagebox.showinfo("Sucesso", f"Arquivo salvo com sucesso em:\n{caminho_saida}")
        else:
            messagebox.showwarning("Cancelado", "Operação cancelada pelo usuário.")

    except PermissionError:
        messagebox.showerror("Permissão negada", "Você deve estar mantendo o arquivo Excel de destino em aberto.")

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao processar o arquivo:\n{e}")
        logg.exception("Erro inesperado")

    stop_progress()

################ Tkinter Básico ################

# Cria janela principal da aplicação
janela = tk.Tk()
janela.geometry("350x175")  # Define tamanho da janela
janela.title("Conversor Pdf powered by Mobi")  # Título da janela

# Título maior e em negrito
tk.Label(janela, text="Conversor de PDF para Excel", font=("Helvetica", 13, "bold")).pack(pady=(10, 15))

# Botão para iniciar a conversão chamando a função processamento
botao = tk.Button(janela,
                  text="Converter PDF",
                  command=processamento,
                  font=("Helvetica", 13, "bold"),
                  height=1,
                  width=15)
botao.pack(pady=1)

# Barra de progresso indeterminada (não mostra porcentagem)
progressbar = ttk.Progressbar(janela, orient='horizontal', length=300, mode='indeterminate')
progressbar.pack_forget()  # Inicialmente oculta a barra

# Executa a interface gráfica
janela.mainloop()
