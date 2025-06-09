#O erro que você está enfrentando está relacionado ao pacote camelot, mais especificamente a uma dependência dele: 
# o SQLAlchemy (em versão muito antiga, 0.7.10), que usa uma biblioteca (lib2to3) que foi removida no Python 3.13.
# Opção 1 (RECOMENDADA): Usar Python 3.10 ou 3.11 camelot e muitos pacotes científicos ainda não são totalmente 
# compatíveis com o Python 3.13.
# modifiquei o Python para 3.11 e mudei o requirements, refiz o ambiente virtual. 
# O executável desse foi criado usando comandos diferentes de execução devido ao fato de usar o Python 3.11, 
# documentados no Python
# Coloquei os comentários via Chat Gpt explicando linha por linha

import pandas as pd  # Biblioteca para manipulação de dados em forma de tabelas (DataFrame)
import tkinter as tk  # Biblioteca para criar interfaces gráficas (GUI)
from tkinter import filedialog, messagebox, ttk  # Componentes de interface do tkinter
import logging  # Biblioteca para registrar logs do sistema (mensagens informativas, erros etc.)
import camelot  # Biblioteca para extração de tabelas de PDFs
import pdfplumber  # Biblioteca para extração de texto de PDFs

from openpyxl import Workbook  # Criação e manipulação de arquivos Excel (.xlsx)
from openpyxl.styles import Border, Side  # Estilização de bordas nas células do Excel
from openpyxl.utils import get_column_letter  # Conversão de número de coluna para letra (ex: 1 -> 'A')
from openpyxl.utils.dataframe import dataframe_to_rows  # Converte DataFrame do pandas em linhas para o Excel

###################### Configuração de Log ######################
logging.basicConfig(level=logging.INFO)  # Define o nível mínimo de log como INFO
logg = logging.getLogger(__name__)  # Cria um logger para o módulo atual

###################### Funções da Barra ######################
def start_progress():
    progressbar.pack(pady=10)  # Exibe a barra de progresso com espaço vertical
    progressbar.start(10)  # Inicia a barra com velocidade (10 ms entre atualizações)
    botao.config(state='disabled')  # Desativa o botão enquanto processa
    janela.update_idletasks()  # Atualiza a interface gráfica imediatamente

def stop_progress():
    progressbar.stop()  # Para a barra de progresso
    progressbar.pack_forget()  # Esconde a barra de progresso
    botao.config(state='normal')  # Reativa o botão após o fim do processo
    janela.update_idletasks()  # Atualiza a interface

###################### Função Principal ######################
def processamento():
    start_progress()  # Inicia a animação da barra de progresso
    logg.info("Processamento iniciado.")  # Registra início do processo

    caminho_arquivo = filedialog.askopenfilename(  # Abre caixa de diálogo para escolher um PDF
        title="Selecione arquivo em PDF:",
        filetypes=[("Arquivo PDF", "*.pdf")]
    )

    if not caminho_arquivo:  # Se o usuário cancelar a seleção
        messagebox.showwarning("Aviso", "Nenhum arquivo foi selecionado.")  # Exibe aviso
        stop_progress()  # Para a barra de progresso
        return  # Interrompe a execução da função

    try:
        ################### Extraindo Tabelas ###################
        logg.info("Lendo tabelas com Camelot...")  # Log de início da leitura

        tables_lattice = camelot.read_pdf(caminho_arquivo, pages='all', flavor='lattice')  # Tenta extrair tabelas com método lattice
        tables_stream = camelot.read_pdf(caminho_arquivo, pages='all', flavor='stream')  # Também tenta com método stream

        if tables_lattice.n >= tables_stream.n and tables_lattice.n != 0:  # Prioriza lattice se for melhor
            tables = tables_lattice
            metodo = 'lattice'
        elif tables_stream.n != 0:  # Usa stream se lattice falhar
            tables = tables_stream
            metodo = 'stream'
        else:  # Nenhuma tabela encontrada
            tables = None
            metodo = None

        logg.info(f"Modo usado para tabelas: {metodo if metodo else 'Nenhuma tabela encontrada'}")  # Log do método escolhido

        if tables:
            dataframe_tabelas = pd.concat([t.df for t in tables])  # Junta todas as tabelas em um único DataFrame
            dataframe_tabelas = dataframe_tabelas.applymap(lambda x: x.strip() if isinstance(x, str) else x)  # Limpa espaços em strings
            dataframe_tabelas = dataframe_tabelas.dropna(how='all', axis=1).dropna(how='all', axis=0)  # Remove colunas e linhas vazias
        else:
            dataframe_tabelas = pd.DataFrame()  # Cria DataFrame vazio se não houver tabela

        ################### Extraindo Texto ###################
        logg.info("Lendo texto com pdfplumber...")  # Log da extração de texto

        texto_completo = ""  # Inicializa variável de texto
        with pdfplumber.open(caminho_arquivo) as pdf:  # Abre PDF
            for pagina in pdf.pages:  # Itera por cada página
                texto_completo += pagina.extract_text() + "\n"  # Extrai texto da página

        if not texto_completo.strip():  # Se não houver texto extraído
            texto_completo = "Nenhum texto encontrado no PDF."  # Mensagem padrão

        ################### Selecionar Local para Salvar ###################
        caminho_saida = filedialog.asksaveasfilename(  # Abre caixa para escolher onde salvar
            defaultextension=".xlsx",  # Define extensão padrão
            filetypes=[("Arquivos Excel", "*.xlsx")],  # Tipos de arquivo permitidos
            title="Salvar arquivo como:"
        )

        if not caminho_saida:  # Se usuário cancelar o salvamento
            messagebox.showwarning("Cancelado", "Operação cancelada pelo usuário.")
            stop_progress()
            return

        ################### Gerando Excel ###################
        logg.info("Gerando arquivo Excel...")  # Log de início do Excel

        wb = Workbook()  # Cria novo arquivo Excel

        # Aba das tabelas
        ws1 = wb.active  # Usa a aba padrão
        ws1.title = "Tabela"  # Renomeia a aba

        if not dataframe_tabelas.empty:  # Se houver tabela
            for r in dataframe_to_rows(dataframe_tabelas, index=False, header=True):  # Converte DataFrame em linhas
                ws1.append(r)  # Adiciona as linhas na planilha

            border = Border(  # Define borda fina para todas as células
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, max_col=ws1.max_column):  # Itera pelas células preenchidas
                for cell in row:
                    if cell.value is not None:  # Se a célula tiver valor
                        cell.border = border  # Aplica borda

        else:
            ws1.append(["Nenhuma tabela encontrada."])  # Texto padrão se não tiver tabela

        # Aba do texto
        ws2 = wb.create_sheet(title="Texto")  # Cria nova aba chamada Texto
        linhas = texto_completo.strip().split("\n")  # Divide o texto em linhas
        for linha in linhas:
            ws2.append([linha])  # Adiciona cada linha como uma nova célula na aba Texto

        wb.save(caminho_saida)  # Salva o arquivo Excel
        logg.info(f"Arquivo Excel salvo com sucesso em: {caminho_saida}")  # Log de sucesso
        messagebox.showinfo("Sucesso", f"Arquivo salvo com sucesso em:\n{caminho_saida}")  # Alerta de sucesso para o usuário

    except Exception as e:  # Se ocorrer erro durante o processo
        logg.exception("Erro durante o processamento.")  # Log do erro
        messagebox.showerror("Erro", f"Ocorreu um erro:\n{e}")  # Alerta do erro

    stop_progress()  # Finaliza a barra de progresso
    logg.info("Processamento finalizado.")  # Log do fim do processo

###################### Interface Tkinter ######################
janela = tk.Tk()  # Cria a janela principal
janela.geometry("350x175")  # Define tamanho da janela
janela.title("Conversor PDF para Excel (Tabelas + Texto)")  # Define o título da janela

tk.Label(janela, text="Conversor PDF para Excel", font=("Helvetica", 13, "bold")).pack(pady=(10, 15))  # Cria título da interface

botao = tk.Button(janela,  # Cria botão principal da interface
                   text="Converter PDF",
                   command=processamento,  # Ao clicar, chama a função principal
                   font=("Helvetica", 13, "bold"),
                   height=1,
                   width=15)
botao.pack(pady=1)  # Adiciona o botão com espaçamento vertical

progressbar = ttk.Progressbar(janela, orient='horizontal', length=300, mode='indeterminate')  # Cria barra de progresso (oculta por padrão)
progressbar.pack_forget()  # Esconde a barra inicialmente

janela.mainloop()  # Inicia o loop da interface (mantém a janela aberta)
