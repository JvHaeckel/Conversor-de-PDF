# Converter PDF para excel usando Python 

import pandas as pd
import tabula
import tkinter as tk 
from tkinter import filedialog, messagebox, ttk

################ Barra de Progresso Funções ################

def start_progress():
    progressbar.pack(pady=10)
    progressbar.start(10)  # Inicia animação da barra indeterminada
    botao.config(state='disabled')
    janela.update_idletasks()

def stop_progress():
    progressbar.stop()
    progressbar.pack_forget()
    botao.config(state='normal')
    janela.update_idletasks()
    
    
################ Pegando Documento da Interface + LENDO O PDF ################

def processamento():
    
    start_progress()

    ################ Pegando Documento da Interface  ################
    
    # Abre janela para o user pegar arquivo em PDF
    caminho_arquivo = filedialog.askopenfilename(  # filedialog: Permite abrir janelas para o usuário selecionar arquivos ou escolher onde salvar.
        title="Selecione arquivo em PDF: ",
        filetypes=[("Arquivo PDF", "*.pdf")]
    )
     
    # Se o usuário não escolher nada exibe esse aviso e encerra o processo 
    if not caminho_arquivo:
         messagebox.showwarning("Aviso", "Nenhum arquivo foi selecionado.") # messagebox: Permite exibir caixas de mensagem (avisos, erros, informações)
         stop_progress()
         return
    
    ################ LENDO O PDF  ################
    try:
        # A função read_pdf do Tabulas lê tabelas do PDF e retorna um DataFrame
        dataframe = tabula.read_pdf(
            caminho_arquivo, 
            pages='all',            # pages='all': Significa que o Tabula deve procurar tabelas em todas as páginas do PDF.
            multiple_tables=True
        )
    
        # Verifica se tem tabelas no PDF
        if not dataframe:
            messagebox.showinfo("Informação", "Nenhuma tabela foi encontrada no PDF.")
            stop_progress()
            return
        
        # Combina todas as tabelas em um único DataFrame usando o pd.concat - é uma função do Pandas para unir DataFrames.
        dataframe_complete = pd.concat(dataframe)   
        
        # Abre a janela para salvar o arquivo, forçando extensão .xlsx
        caminho_saida = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx")],
            title="Salvar arquivo como:"
        )
        
        ###################### Tratamento de Linhas e colunas vazias ######################
        
        dataframe_complete = dataframe_complete.applymap(lambda x: x.strip() if isinstance(x, str) else x)  # Remove espaços em branco extras de todas as células (aplica para strings)
        dataframe_complete = dataframe_complete.dropna(how='all', axis=1)  # Remove colunas vazias
        dataframe_complete = dataframe_complete.dropna(how='all', axis=0)  # Remove linhas vazias
        
        ###################### Exportando com openpyxl formatado ######################
        if caminho_saida:
            from openpyxl import Workbook
            from openpyxl.styles import Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.utils.dataframe import dataframe_to_rows

            # Cria uma nova planilha
            wb = Workbook()
            ws = wb.active
            ws.title = "PDF_Convertido"

            # Adiciona os dados do DataFrame na planilha
            for r in dataframe_to_rows(dataframe_complete, index=False, header=True):
                ws.append(r)

            # Define borda fina para aplicar nas células com dados
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Aplica borda nas células com conteúdo
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    if cell.value is not None:
                        cell.border = border

            # Oculta colunas extras após os dados (visual)
            for col in range(ws.max_column + 1, 50):
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].hidden = True

            # Oculta linhas extras após os dados (visual)
            for row in range(ws.max_row + 1, 200):
                ws.row_dimensions[row].hidden = True

            # Salva o arquivo formatado
            wb.save(caminho_saida)
            messagebox.showinfo("Sucesso", f"Arquivo salvo com sucesso em:\n{caminho_saida}")
        
        else:
            messagebox.showwarning("Cancelado", "Operação cancelada pelo usuário.")
    
    # Tratando erro de deixar arquivo do excel de destino em aberto.
    except PermissionError:
        messagebox.showerror("Permissão negada", "Você deve estar mantendo o arquivo Excel de destino em aberto.")    
        
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao processar o arquivo:\n{e}")

    stop_progress()

################ Tkinter Básico ################

janela = tk.Tk()
janela.geometry("350x175")
janela.title("Conversor Pdf powered by Mobi")

# Título maior e em negrito
tk.Label(janela, text="Conversor de PDF para Excel", font=("Helvetica", 13, "bold")).pack(pady=(10, 15))

botao = tk.Button(janela, 
                  text="Converter PDF", 
                  command=processamento, 
                  font=("Helvetica", 13, "bold"),
                  height=1, 
                  width=15)
botao.pack(pady=1)

progressbar = ttk.Progressbar(janela, orient='horizontal', length=300, mode='indeterminate')
progressbar.pack_forget()

janela.mainloop()
