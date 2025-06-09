#O erro que você está enfrentando está relacionado ao pacote camelot, mais especificamente a uma dependência dele: 
# o SQLAlchemy (em versão muito antiga, 0.7.10), que usa uma biblioteca (lib2to3) que foi removida no Python 3.13.
# Opção 1 (RECOMENDADA): Usar Python 3.10 ou 3.11 camelot e muitos pacotes científicos ainda não são totalmente 
# compatíveis com o Python 3.13.
# modifiquei o Python para 3.11 e mudei o requirements, refiz o ambiente virtual. 
# O executável desse foi criado usando comandos diferentes de execução devido ao fato de usar o Python 3.11, 
# documentados no Python



import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import logging
import camelot
import pdfplumber

from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

###################### Configuração de Log ######################
logging.basicConfig(level=logging.INFO)
logg = logging.getLogger(__name__)

###################### Funções da Barra ######################
def start_progress():
    progressbar.pack(pady=10)
    progressbar.start(10)
    botao.config(state='disabled')
    janela.update_idletasks()

def stop_progress():
    progressbar.stop()
    progressbar.pack_forget()
    botao.config(state='normal')
    janela.update_idletasks()

###################### Função Principal ######################
def processamento():
    start_progress()
    logg.info("Processamento iniciado.")

    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione arquivo em PDF:",
        filetypes=[("Arquivo PDF", "*.pdf")]
    )

    if not caminho_arquivo:
        messagebox.showwarning("Aviso", "Nenhum arquivo foi selecionado.")
        stop_progress()
        return

    try:
        ################### Extraindo Tabelas ###################
        logg.info("Lendo tabelas com Camelot...")

        tables_lattice = camelot.read_pdf(caminho_arquivo, pages='all', flavor='lattice')
        tables_stream = camelot.read_pdf(caminho_arquivo, pages='all', flavor='stream')

        if tables_lattice.n >= tables_stream.n and tables_lattice.n != 0:
            tables = tables_lattice
            metodo = 'lattice'
        elif tables_stream.n != 0:
            tables = tables_stream
            metodo = 'stream'
        else:
            tables = None
            metodo = None

        logg.info(f"Modo usado para tabelas: {metodo if metodo else 'Nenhuma tabela encontrada'}")

        if tables:
            dataframe_tabelas = pd.concat([t.df for t in tables])
            dataframe_tabelas = dataframe_tabelas.applymap(lambda x: x.strip() if isinstance(x, str) else x)
            dataframe_tabelas = dataframe_tabelas.dropna(how='all', axis=1).dropna(how='all', axis=0)
        else:
            dataframe_tabelas = pd.DataFrame()

        ################### Extraindo Texto ###################
        logg.info("Lendo texto com pdfplumber...")

        texto_completo = ""
        with pdfplumber.open(caminho_arquivo) as pdf:
            for pagina in pdf.pages:
                texto_completo += pagina.extract_text() + "\n"

        if not texto_completo.strip():
            texto_completo = "Nenhum texto encontrado no PDF."

        ################### Selecionar Local para Salvar ###################
        caminho_saida = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx")],
            title="Salvar arquivo como:"
        )

        if not caminho_saida:
            messagebox.showwarning("Cancelado", "Operação cancelada pelo usuário.")
            stop_progress()
            return

        ################### Gerando Excel ###################
        logg.info("Gerando arquivo Excel...")

        wb = Workbook()

        # Aba das tabelas
        ws1 = wb.active
        ws1.title = "Tabela"

        if not dataframe_tabelas.empty:
            for r in dataframe_to_rows(dataframe_tabelas, index=False, header=True):
                ws1.append(r)

            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, max_col=ws1.max_column):
                for cell in row:
                    if cell.value is not None:
                        cell.border = border

        else:
            ws1.append(["Nenhuma tabela encontrada."])

        # Aba do texto
        ws2 = wb.create_sheet(title="Texto")
        linhas = texto_completo.strip().split("\n")
        for linha in linhas:
            ws2.append([linha])

        wb.save(caminho_saida)
        logg.info(f"Arquivo Excel salvo com sucesso em: {caminho_saida}")
        messagebox.showinfo("Sucesso", f"Arquivo salvo com sucesso em:\n{caminho_saida}")

    except Exception as e:
        logg.exception("Erro durante o processamento.")
        messagebox.showerror("Erro", f"Ocorreu um erro:\n{e}")

    stop_progress()
    logg.info("Processamento finalizado.")

###################### Interface Tkinter ######################
janela = tk.Tk()
janela.geometry("350x175")
janela.title("Conversor PDF para Excel (Tabelas + Texto)")

tk.Label(janela, text="Conversor PDF para Excel", font=("Helvetica", 13, "bold")).pack(pady=(10, 15))

botao = tk.Button(janela,
                   text="Converter PDF",
                   command=processamento,
                   font=("Helvetica", 13, "bold"),
                   height=1,
                   width=15)
botao.pack(pady=1)

progressbar = ttk.Progressbar(janela, orient='horizontal', length=300, mode='indeterminate')
progressbar.pack_forget()

janela.mainloop()
